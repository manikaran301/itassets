import openpyxl
import json
import os

excel_path = os.path.join(os.path.dirname(__file__), '..', 'public', 'Designation.xlsx')
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

designations = []
# Row 1 is header ("Designation"), so we start from Row 2
for row_idx in range(2, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=1).value
    if val:
        cleaned_val = str(val).strip()
        if cleaned_val and cleaned_val not in designations:
            designations.append(cleaned_val)

print(json.dumps(designations))
