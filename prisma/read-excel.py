import openpyxl
import json
import os

excel_path = os.path.join(os.path.dirname(__file__), '..', 'public', 'CURRENT EMPLOYEES.xlsx')
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

rows = []
for row_idx in range(2, ws.max_row + 1):
    emp_id = ws.cell(row=row_idx, column=2).value
    name = ws.cell(row=row_idx, column=3).value
    company = ws.cell(row=row_idx, column=4).value
    location = ws.cell(row=row_idx, column=5).value
    if emp_id and name:
        rows.append({
            "empId": str(emp_id).strip(),
            "fullName": str(name).strip(),
            "company": str(company).strip() if company else "",
            "location": str(location).strip() if location else ""
        })

print(json.dumps(rows))
